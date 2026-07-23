"""Build catalog.db from products_detail.xlsx.

Everything expensive happens here, once, offline: HTML unescaping, number
parsing, EAV spec rows flattened into canonical numeric features, the full-text
index, and per-category price distributions. At request time the server only
ever runs indexed SQL, which is what keeps the latency budget intact.

    python backend/ingest.py [--xlsx PATH] [--db PATH]
"""

from __future__ import annotations

import argparse
import html
import sqlite3
import sys
import time
from pathlib import Path

import openpyxl

sys.path.insert(0, str(Path(__file__).resolve().parent))

from taxonomy import category_group, extract_features  # noqa: E402
from vntext import fold, parse_sold_count, parse_spec_value  # noqa: E402

ROOT = Path(__file__).resolve().parent.parent
DEFAULT_XLSX = ROOT / "data" / "products_detail.xlsx"
DEFAULT_DB = ROOT / "data" / "catalog.db"

SCHEMA = """
PRAGMA journal_mode = WAL;

DROP TABLE IF EXISTS products;
CREATE TABLE products (
    product_id    TEXT PRIMARY KEY,
    name          TEXT NOT NULL,
    name_fold     TEXT NOT NULL,
    category      TEXT NOT NULL,
    category_fold TEXT NOT NULL,
    cat_group     TEXT NOT NULL,
    brand         TEXT,
    brand_fold    TEXT,
    price         INTEGER,          -- Giá khuyến mãi: what the customer pays
    price_list    INTEGER,          -- Giá gốc
    discount_pct  REAL,
    rating        REAL,
    n_sold        INTEGER,
    color         TEXT,
    online_only   INTEGER,
    warranty      TEXT,
    promotion     TEXT,
    has_promo     INTEGER,
    url           TEXT,
    image_url     TEXT
);

DROP TABLE IF EXISTS specs;
CREATE TABLE specs (
    product_id TEXT NOT NULL,
    key        TEXT NOT NULL,
    key_fold   TEXT NOT NULL,
    value      TEXT,
    value_fold TEXT
);

DROP TABLE IF EXISTS features;
CREATE TABLE features (
    product_id TEXT NOT NULL,
    feat       TEXT NOT NULL,
    value      REAL NOT NULL
);

DROP TABLE IF EXISTS category_stats;
CREATE TABLE category_stats (
    cat_group TEXT PRIMARY KEY,
    n         INTEGER,
    p10 INTEGER, p25 INTEGER, p50 INTEGER, p75 INTEGER, p90 INTEGER
);

-- Per-group distribution of every numeric feature. Lets slang be expressed
-- relatively ("pin trâu" = top 40% battery for *this* category) instead of with
-- hardcoded constants that would be wrong for all but one category.
DROP TABLE IF EXISTS feature_stats;
CREATE TABLE feature_stats (
    cat_group TEXT NOT NULL,
    feat      TEXT NOT NULL,
    n         INTEGER,
    p10 REAL, p25 REAL, p40 REAL, p50 REAL, p60 REAL, p75 REAL, p90 REAL,
    PRIMARY KEY (cat_group, feat)
);

DROP TABLE IF EXISTS products_fts;
CREATE VIRTUAL TABLE products_fts USING fts5(
    product_id UNINDEXED,
    text,
    tokenize = "unicode61 remove_diacritics 2"
);
"""

INDEXES = """
CREATE INDEX idx_prod_group  ON products(cat_group, price);
CREATE INDEX idx_prod_price  ON products(price);
CREATE INDEX idx_prod_brand  ON products(brand_fold);
CREATE INDEX idx_prod_cat    ON products(category_fold);
CREATE INDEX idx_specs_pid   ON specs(product_id);
CREATE INDEX idx_specs_key   ON specs(key_fold);
CREATE INDEX idx_feat_pid    ON features(product_id);
CREATE INDEX idx_feat_lookup ON features(feat, value);
"""


def clean(v: object) -> str | None:
    """Trim and HTML-unescape a cell.

    The crawl left entities raw — 'Lock&amp;Lock', 'AVA&#x2B;' — so brand
    matching fails on the literal text unless this runs first.
    """
    if v is None:
        return None
    s = html.unescape(str(v)).strip()
    return s or None


def as_int(v: object) -> int | None:
    if isinstance(v, bool):
        return int(v)
    if isinstance(v, (int, float)):
        return int(v)
    if isinstance(v, str):
        digits = "".join(c for c in v if c.isdigit())
        return int(digits) if digits else None
    return None


def as_float(v: object) -> float | None:
    if isinstance(v, (int, float)) and not isinstance(v, bool):
        return float(v)
    if isinstance(v, str):
        try:
            return float(v.replace(",", ".").strip())
        except ValueError:
            return None
    return None


def percentile(sorted_vals: list[int], p: float) -> int:
    if not sorted_vals:
        return 0
    idx = min(len(sorted_vals) - 1, max(0, int(round(p * (len(sorted_vals) - 1)))))
    return sorted_vals[idx]


def load_products(ws, con: sqlite3.Connection, fts_batch: dict[str, str]) -> dict[str, str]:
    rows = ws.iter_rows(values_only=True)
    hdr = [str(h) for h in next(rows)]
    i = {h: n for n, h in enumerate(hdr)}

    group_of: dict[str, str] = {}
    prices: dict[str, list[int]] = {}
    batch: list[tuple] = []
    seen: set[str] = set()
    dupes = 0

    for r in rows:
        pid = clean(r[i["product_id"]])
        if not pid:
            continue
        if pid in seen:
            dupes += 1
            continue
        seen.add(pid)

        name = clean(r[i["tên sản phẩm"]]) or ""
        category = clean(r[i["category_name"]]) or ""
        brand = clean(r[i["brand"]])
        cat_fold = fold(category)
        group = category_group(cat_fold)
        group_of[pid] = group

        price = as_int(r[i["Giá khuyến mãi"]])
        price_list = as_int(r[i["Giá gốc"]])
        # A handful of rows carry no promo price; fall back to list price so the
        # product is still reachable instead of silently dropping out of range
        # filters as a NULL.
        if not price:
            price = price_list
        discount = None
        if price and price_list and price_list > price:
            discount = round(1.0 - price / price_list, 4)

        promotion = clean(r[i["promotion"]])
        batch.append((
            pid, name, fold(name), category, cat_fold, group,
            brand, fold(brand), price, price_list, discount,
            as_float(r[i["rating_vote"]]), parse_sold_count(r[i["quantity_sold"]]),
            clean(r[i["màu sắc"]]), 1 if r[i["onlineSaleOnly"]] else 0,
            clean(r[i["chính sách bảo hành"]]), promotion, 1 if promotion else 0,
            clean(r[i["url"]]), clean(r[i["url_image"]]),
        ))
        # Index the *folded* text. SQLite's "remove_diacritics 2" strips
        # combining marks but leaves đ (U+0111) alone — it is precomposed with a
        # stroke and never decomposes. Indexing raw text means "điện thoại" is
        # tokenised as "đien thoai" and a folded query for "dien thoai" matches
        # nothing at all. Folding both sides keeps them in agreement.
        fts_batch[pid] = fold(" ".join(filter(None, [
            name, brand, category, clean(r[i["màu sắc"]]), promotion,
        ])))
        if price:
            prices.setdefault(group, []).append(price)

    con.executemany(
        "INSERT INTO products VALUES (" + ",".join("?" * 20) + ")", batch)

    stats = []
    for group, vals in prices.items():
        vals.sort()
        stats.append((group, len(vals), *(percentile(vals, p)
                                          for p in (0.10, 0.25, 0.50, 0.75, 0.90))))
    con.executemany("INSERT INTO category_stats VALUES (?,?,?,?,?,?,?)", stats)

    print(f"  products : {len(batch):,} rows"
          + (f"  ({dupes} duplicate ids skipped)" if dupes else ""))
    print(f"  groups   : {len(prices)} coarse groups with price data")
    return group_of


def load_specs(ws, con: sqlite3.Connection, known: set[str],
               fts_batch: dict[str, str]) -> None:
    rows = ws.iter_rows(values_only=True)
    next(rows)  # header

    spec_batch, feat_batch = [], []
    n_rows = orphans = 0
    # Same (product, feature) can be described by several spec rows; keep the
    # largest, which is the headline figure for every feature we extract.
    best: dict[tuple[str, str], float] = {}

    for r in rows:
        pid = clean(r[0])
        key = clean(r[2])
        if not pid or not key:
            continue
        if pid not in known:
            orphans += 1
            continue
        n_rows += 1
        value = clean(r[3])
        key_f = fold(key)
        value_f = fold(value)
        spec_batch.append((pid, key, key_f, value, value_f))
        # Fold spec text into the search index too, so free-text constraints
        # like "inverter", "IPX5" or "RTX" are one MATCH away instead of an
        # EXISTS scan over 212k spec rows per candidate.
        if value_f:
            fts_batch[pid] = f"{fts_batch.get(pid, '')} {value_f}"

        num, unit = parse_spec_value(value)
        hit = extract_features(key_f, value, num, unit)
        if hit:
            feat, val = hit
            k = (pid, feat)
            if val > best.get(k, float("-inf")):
                best[k] = val

        if len(spec_batch) >= 50_000:
            con.executemany("INSERT INTO specs VALUES (?,?,?,?,?)", spec_batch)
            spec_batch.clear()

    if spec_batch:
        con.executemany("INSERT INTO specs VALUES (?,?,?,?,?)", spec_batch)

    feat_batch = [(pid, feat, val) for (pid, feat), val in best.items()]
    con.executemany("INSERT INTO features VALUES (?,?,?)", feat_batch)

    print(f"  specs    : {n_rows:,} rows"
          + (f"  ({orphans} orphaned product ids skipped)" if orphans else ""))
    print(f"  features : {len(feat_batch):,} numeric values extracted")


def build_feature_stats(con: sqlite3.Connection) -> None:
    """Percentile table per (category group, feature).

    Needs a meaningful sample to be trustworthy — a percentile over three rows
    describes nothing — so groups with too few values are skipped and the
    lexicon falls back to an absolute threshold for them.
    """
    MIN_SAMPLE = 8
    buckets: dict[tuple[str, str], list[float]] = {}
    for group, feat, val in con.execute(
        "SELECT p.cat_group, f.feat, f.value FROM features f "
        "JOIN products p USING(product_id)"
    ):
        buckets.setdefault((group, feat), []).append(val)

    rows = []
    for (group, feat), vals in buckets.items():
        if len(vals) < MIN_SAMPLE:
            continue
        vals.sort()
        def pct(p: float) -> float:
            idx = min(len(vals) - 1, max(0, int(round(p * (len(vals) - 1)))))
            return vals[idx]
        rows.append((group, feat, len(vals), pct(.10), pct(.25), pct(.40),
                     pct(.50), pct(.60), pct(.75), pct(.90)))

    con.executemany(
        "INSERT INTO feature_stats VALUES (?,?,?,?,?,?,?,?,?,?)", rows)
    print(f"  featstats: {len(rows):,} (group, feature) distributions")


def main() -> int:
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("--xlsx", type=Path, default=DEFAULT_XLSX)
    ap.add_argument("--db", type=Path, default=DEFAULT_DB)
    args = ap.parse_args()

    if not args.xlsx.exists():
        print(f"error: {args.xlsx} not found", file=sys.stderr)
        return 1

    t0 = time.time()
    print(f"reading {args.xlsx.name} ...")
    wb = openpyxl.load_workbook(args.xlsx, read_only=True, data_only=True)

    args.db.parent.mkdir(parents=True, exist_ok=True)
    con = sqlite3.connect(args.db)
    con.executescript(SCHEMA)

    fts_batch: dict[str, str] = {}
    group_of = load_products(wb["products"], con, fts_batch)
    load_specs(wb["specs"], con, set(group_of), fts_batch)
    con.executemany("INSERT INTO products_fts VALUES (?,?)", fts_batch.items())
    build_feature_stats(con)

    con.executescript(INDEXES)
    con.execute("INSERT INTO products_fts(products_fts) VALUES('optimize')")
    con.commit()

    n_feats = con.execute(
        "SELECT COUNT(DISTINCT feat) FROM features").fetchone()[0]
    con.execute("ANALYZE")
    con.commit()
    con.close()
    wb.close()

    size_mb = args.db.stat().st_size / 1e6
    print(f"\nwrote {args.db}  ({size_mb:.1f} MB, {n_feats} distinct features)")
    print(f"done in {time.time() - t0:.1f}s")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
